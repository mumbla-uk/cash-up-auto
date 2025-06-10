import imapclient
import ssl
import re
import openpyxl
from datetime import datetime
import email
from bs4 import BeautifulSoup
import os
import streamlit as st

# --- Configuration ---
# Retrieve secrets from Streamlit's secrets management
GMAIL_USER = st.secrets["GMAIL_USER"]
GMAIL_APP_PASSWORD = st.secrets["GMAIL_APP_PASSWORD"]
IMAP_SERVER = 'imap.gmail.com'
IMAP_PORT = 993
EMAIL_SUBJECT = st.secrets["EMAIL_SUBJECT"]

# Define the Excel file path relative to the script's directory, or an absolute path
EXCEL_FILE_NAME = 'cash_up_records.xlsx'
EXCEL_FILE_PATH = os.path.join(os.path.dirname(__file__), EXCEL_FILE_NAME)


# --- Excel Headers ---
EXCEL_HEADERS = [
    'Date', 'Z', 'Card', 'Cash', 'Petty Cash',
    'Card Gratuity', 'Cash Gratuity', 'Total Gratuity',
    'Discounts', 'Difference'
]

def get_email_body(msg):
    """
    Extracts the plain text or HTML body from an email message.
    Prioritizes plain text if available, otherwise attempts to parse HTML.
    """
    if msg.is_multipart():
        for part in msg.walk():
            ctype = part.get_content_type()
            cdispo = str(part.get('Content-Disposition'))

            if ctype == 'text/plain' and 'attachment' not in cdispo:
                return part.get_payload(decode=True).decode('utf-8')
            if ctype == 'text/html' and 'attachment' not in cdispo:
                html_content = part.get_payload(decode=True).decode('utf-8')
                soup = BeautifulSoup(html_content, 'html.parser')
                return soup.get_text(separator='\n')
    else:
        ctype = msg.get_content_type()
        if ctype == 'text/plain':
            return msg.get_payload(decode=True).decode('utf-8')
        elif ctype == 'text/html':
            html_content = msg.get_payload(decode=True).decode('utf-8')
            soup = BeautifulSoup(html_content, 'html.parser')
            return soup.get_text(separator='\n')
    return None

def parse_cash_up_summary(text):
    """
    Parses the text summary from the email with the 'Key = Value' format.
    """
    data = {}
    lines = text.strip().split('\n')

    for line in lines:
        line = line.strip()
        if '=' in line:
            try:
                key, value_str = line.split('=', 1)
                key = key.strip()
                value_str = value_str.strip()

                if key == 'Date':
                    try:
                        parsed_date = datetime.strptime(value_str, '%A, %B %d, %Y').strftime('%Y-%m-%d')
                        data[key] = parsed_date
                    except ValueError:
                        # Fallback for date parsing if specific format fails, try common YYYY-MM-DD
                        try:
                            datetime.strptime(value_str, '%Y-%m-%d') # Just to validate format
                            data[key] = value_str
                        except ValueError:
                            print(f"Warning: Could not parse date format '{value_str}'. Storing as is.")
                            data[key] = value_str
                else:
                    value_str = value_str.replace('£', '').strip()
                    data[key] = float(value_str)
            except ValueError as e:
                print(f"Warning: Could not parse value from line '{line}'. Error: {e}")
                continue

    for header in EXCEL_HEADERS:
        if header not in data:
            if header == 'Date':
                data[header] = ''
            else:
                data[header] = 0.0

    return data

def add_data_to_excel(excel_path, data_row, headers):
    """
    Appends a new row of data to the correct monthly sheet in the Excel file.
    Creates the file and sheets with headers if they don't exist.
    """
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        # Remove the default sheet created on new workbook, if it's empty
        if 'Sheet' in workbook.sheetnames:
            std_sheet = workbook['Sheet']
            if len(std_sheet._cells) == 0: # Check if sheet is actually empty
                workbook.remove(std_sheet)


    # Determine the sheet name based on the date
    date_str = data_row.get('Date')
    if not date_str:
        sheet_name = "Undated Entries" # Fallback if no date is found
    else:
        try:
            # Assuming date_str is 'YYYY-MM-DD'
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            sheet_name = date_obj.strftime('%Y-%m') # e.g., "2025-06"
        except ValueError:
            sheet_name = "Invalid Date Entries" # Fallback for unparseable dates

    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(headers) # Add headers to the new sheet
    else:
        sheet = workbook[sheet_name]

    ordered_row_values = []
    for header in headers:
        ordered_row_values.append(data_row.get(header, 0.0 if header != 'Date' else ''))

    sheet.append(ordered_row_values)
    workbook.save(excel_path)

def run_automation():
    status_messages = []
    try:
        context = ssl.create_default_context()
        with imapclient.IMAPClient(IMAP_SERVER, port=IMAP_PORT, ssl_context=context) as client:
            client.login(GMAIL_USER, GMAIL_APP_PASSWORD)
            status_messages.append("Logged in to Gmail successfully.")

            # Select 'INBOX' and search for UNSEEN messages
            # This is crucial for processing unique emails only
            client.select_folder('INBOX')
            messages = client.search(['UNSEEN', 'FROM', CASH_UP_SENDER_EMAIL, 'SUBJECT', EMAIL_SUBJECT])

            if not messages:
                status_messages.append(f"No *UNSEEN* emails found with subject '{EMAIL_SUBJECT}' from '{CASH_UP_SENDER_EMAIL}'.")
                status_messages.append("This means all matching emails have already been processed.")
                status_messages.append("If you want to re-process, you'll need to manually mark them as unread in Gmail.")
                return "No New Emails", status_messages # Changed status for clarity

            # Process the latest UNSEEN email
            status_messages.append(f"Found {len(messages)} UNSEEN email(s). Processing the latest one.")
            raw_message = client.fetch(messages[-1], ['BODY[]', 'RFC822'])
            msg_bytes = raw_message[messages[-1]][b'RFC822']
            msg = email.message_from_bytes(msg_bytes)

            email_body = get_email_body(msg)

            if email_body:
                status_messages.append("Email content extracted successfully.")
                parsed_data = parse_cash_up_summary(email_body)
                status_messages.append(f"Email data parsed: {parsed_data.get('Date', 'N/A')} - Z: {parsed_data.get('Z', 'N/A')}")

                add_data_to_excel(EXCEL_FILE_PATH, parsed_data, EXCEL_HEADERS)
                status_messages.append(f"Data added to {EXCEL_FILE_NAME} in sheet: {datetime.strptime(parsed_data.get('Date', ''), '%Y-%m-%d').strftime('%Y-%m') if parsed_data.get('Date') else 'N/A'}")

                # Mark the processed email as 'seen' (read) so it's not processed again
                client.add_flags(messages[-1], ['\Seen'])
                status_messages.append("Email marked as seen in Gmail.")

                return "Success", status_messages
            else:
                status_messages.append("Could not extract email body from the message.")
                return "Error", status_messages

    except imapclient.exceptions.LoginError:
        status_messages.append("Login failed! Please check your Gmail address and App Password.")
        status_messages.append("Ensure you have generated an App Password and enabled 2-Step Verification.")
        return "Login Failed", status_messages
    except Exception as e:
        status_messages.append(f"An unexpected error occurred: {e}")
        # import traceback # Uncomment for full traceback in console for debugging
        # traceback.print_exc()
        return "Error", status_messages
